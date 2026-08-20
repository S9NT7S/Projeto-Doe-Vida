from flask import render_template, request, redirect, url_for, session, jsonify, send_file, flash
from flask_login import current_user, login_required
from controllers.base_controller import BaseController
from banco.BancoMySQL import BancoMySQL
from services import usuario_service
from repositories import usuario_repository
from fpdf import FPDF
from openpyxl.styles import Font, Border, Side, Alignment
from io import BytesIO
import pandas
from pandas import *
from openpyxl import *

class AdminUsuarioController(BaseController):
    def __init__(self, app, usuario_service, grupo_service):
        self.usuario_service = usuario_service
        self.grupo_service = grupo_service  
        self.rotas = [
            ('/listar_usuarios', 'listar_usuarios', self.proteger_rota(self.listar_usuarios)),
            ('/novo_usuario', 'novo_usuario', self.proteger_rota(self.novo_usuario), ['GET', 'POST']),
            ('/admin/usuarios/editar/<int:usuario_id>', 'editar_usuario', self.proteger_rota(self.editar_usuario), ['GET', 'POST']),
            ('/admin/usuarios/excluir/<int:usuario_id>', 'excluir_usuario', self.proteger_rota(self.excluir_usuario), ['POST']),
            ('/exportar_excel', 'exportar_excel', self.proteger_rota(self.exportar_excel)),
            ('/exportar_pdf', 'exportar_pdf', self.proteger_rota(self.exportar_pdf)),
        ]
        super().__init__(app)

        self.database = BancoMySQL()

    def _somente_admin(self):
        return session.get("perfil_logado") == "admin"

    def listar_usuarios(self):
        if not self._somente_admin():
            return redirect(url_for("home"))
        usuarios = self.usuario_service.obter_todos_usuarios()  
        return render_template("listar_usuarios.html", usuarios=usuarios, admin_id=session.get("usuario_id"))

    def novo_usuario(self): 
        if not self._somente_admin():
            return redirect(url_for("home"))

        if request.method == 'POST':
            nome = request.form.get("nome")
            email = request.form.get("email")
            senha = request.form.get("senha")
            perfil = request.form.get("perfil")
            sexo = request.form.get("sexo")
            sangue = request.form.get("sangue")
            idade = request.form.get("idade")

            try:
           
                self.usuario_service.att_user(
                    nome=nome,
                    email=email,
                    senha=senha,
                    perfil=perfil,
                    sexo=sexo,
                    sangue=sangue,
                    idade=idade
                )
               
                return redirect(url_for("listar_usuarios"))

            except ValueError as e:
                
                erro = str(e)
                return render_template(
                    "novo_usuario.html",
                    email=email,
                    perfil=perfil,
                    erro=erro
                )

            except Exception as e:
           
                erro = f"Erro ao cadastrar usuário: {str(e)}"
                return render_template(
                    "novo_usuario.html",
                    email=email,
                    perfil=perfil,
                    erro=erro
                )

        return render_template("novo_usuario.html")

    def excluir_usuario(self, usuario_id):
        # if not self._somente_admin():
        #     return redirect(url_for("home"))

        if session.get("usuario_id") == usuario_id:
            flash("Você não pode excluir sua própria conta.", "erro")
            return redirect(url_for("listar_usuarios"))
        
        try:
            if self.usuario_service.excluir_usuario(usuario_id) == session.get("usuario_id") == usuario_id:
                print("Você não pode excluir o admin")
                raise ValueError
            else:
                self.usuario_service.excluir_usuario(usuario_id)
                flash("Usuário excluído com sucesso.", "sucesso")
        except ValueError:
            flash(str(ValueError), "erro")
        
        return redirect(url_for("listar_usuarios"))

    def editar_usuario(self, usuario_id): 
        if not self._somente_admin():
            return redirect(url_for("home"))

        usuario = self.usuario_service.obter_usuario_por_id(usuario_id)

        print(usuario)

        if request.method == 'POST':
            
            novo_nome = request.form.get("nome")
            nova_senha = request.form.get("senha")
            novo_perfil = request.form.get("perfil")
            novo_sexo = request.form.get("sexo")
            novo_sangue = request.form.get("sangue")
            nova_idade = request.form.get("idade")

            try:

                self.usuario_service.att_user(usuario_id, novo_nome, nova_senha, novo_perfil, novo_sexo, novo_sangue, nova_idade)

                return redirect(url_for("listar_usuarios"))
            
            except ValueError as e:
                erro = str(e)
                return render_template(
                    "editar_usuario.html",
                    usuario={
                        "id": usuario_id
                    },
                    erro=erro
                )

            except Exception as e:
                erro = f"Erro ao atualizar usuário: {str(e)}"
                return render_template(
                    "editar_usuario.html",
                    usuario={
                        "id": usuario_id,
                    },
                    erro=erro
                )
        return render_template(
            "editar_usuario.html",
            usuario=usuario,
        )
    
    def exportar_excel(self):
        if not self._somente_admin():
            return redirect (url_for("home"))
        
        usuarios = self.usuario_service.obter_todos_usuarios()

        dataframe = pandas.DataFrame(usuarios)
        dataframe = dataframe[['nome', 'usuario', 'perfil', 'sexo', 'sangue', 'idade']]

        dataframe.columns = ['nome', 'usuario', 'perfil', 'sexo', 'sangue', 'idade']

        output = BytesIO()
        with pandas.ExcelWriter(output, engine='openpyxl') as writer:
            dataframe.to_excel(writer, index=False, sheet_name='Usuários')

            workbook = writer.book
            worksheet = writer.sheets['Usuários']

            total_usuarios = len(dataframe)
            worksheet.cell(row=1, column=7, value=f"Total de usuários = {total_usuarios}")

            medium_border = Border(
                
                left=Side(style='medium'),
               
                right=Side(style='medium'),
              
                top=Side(style='medium'),
              
                bottom=Side(style='medium')
            )

            for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
                for cell in row:
                    cell.border = medium_border
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    if cell.row == 1:
                        cell.font = Font(bold=True)

        output.seek(0)
        return send_file(
            output,
            download_name="usuarios.xlsx",
            as_attachment=True,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    
    def exportar_pdf(self):
        if not self._somente_admin():
            return redirect(url_for("home"))
        
        usuarios = self.usuario_service.obter_todos_usuarios()
        total = len(usuarios)

        pdf = FPDF()
        pdf.add_page()
        pdf.set_font("Courier", "B", 14)
        pdf.cell(0, 10, "Lista de Usuários", ln=True, align="C")
        pdf.ln(5)
        pdf.set_font("Courier", "B", 10)

        pdf.cell(90, 8, "Id", border=1, align="C")
        pdf.cell(90, 8, "Usuário", border=1, align="C")
        pdf.cell(90, 8, "Perfil", border=1, align="C")
        pdf.set_font("Courier", "", 10)
        for i in usuarios:
            pdf.cell(90, 8, str(i["nome"]), border=1, align="C")
            pdf.cell(90, 8, str(i["usuario"]), border=1, align="C")
            pdf.cell(90, 8, str(i["perfil"]), border=1, ln=True, align="C")
            pdf.cell(90, 8, str(i["sangue"]), border=1, align="C")
            pdf.cell(90, 8, str(i["sexo"]), border=1, align="C")
            pdf.cell(90, 8, str(i["idade"]), border=1, align="C")

        pdf.ln(5)
        pdf.set_font("Courier", "B", 10)
        pdf.cell(0, 10, f"Total de usuários: {total}", ln=True)
        output = BytesIO()
        pdf_bytes = pdf.output(dest="S").encode('latin1', errors='replace')
        output.write(pdf_bytes)
        output.seek(0)

        return send_file(
            output,
            download_name="usuarios.pdf",
            as_attachment=True,
            mimetype="application/pdf"
        )
    
    